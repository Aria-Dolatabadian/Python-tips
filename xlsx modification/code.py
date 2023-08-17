#This code reads xlsx file and highlights different RGAs with different colour also highlights values (gene size) between 1000000 and 5000000

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Read the XLSX file
xlsx_file = 'data.xlsx'
df = pd.read_excel(xlsx_file)
# Get unique gene types
gene_types = df['Type'].unique()
# Define colors for each gene type
colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', 'FF00FF', '00FFFF', 'FFA500']
# Load the XLSX file with openpyxl
book = load_workbook(xlsx_file)
writer = pd.ExcelWriter('modified_data.xlsx', engine='openpyxl')
writer.book = book
# Get the default sheet name (Assuming only one sheet is present)
sheet_name = writer.book.sheetnames[0]
worksheet = writer.book[sheet_name]
# Iterate over gene types and color the cells
for gene_type, color in zip(gene_types, colors):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for row in range(2, df.shape[0] + 2):  # Assuming the data starts from row 2
        cell = f'F{row}'  # Assuming the 'Type' column is at column F
        if df.at[row - 2, 'Type'] == gene_type:
            worksheet[cell].fill = fill
# Iterate over rows and highlight cells in the 'Size' column based on the condition
for row in range(2, df.shape[0] + 2):  # Assuming the data starts from row 2
    cell = f'E{row}'  # Assuming the 'Size' column is at column E
    cell_value = df.at[row - 2, 'Size']
    if 1000000 <= cell_value <= 5000000:
        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet[cell].fill = fill
# Save the updated XLSX file
writer.save()
writer.close()
print("Data exported successfully.")
