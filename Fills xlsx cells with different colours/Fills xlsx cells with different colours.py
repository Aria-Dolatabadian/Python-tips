import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the CSV file
df = pd.read_csv('gene_data.csv')

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set up colors for each gene type
gene_types = df['Type'].unique()
colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', '00FFFF', 'FF00FF', '800000', '008000', '000080', '808000']

# Assign a color to each gene type
color_map = {}
for i, gene_type in enumerate(gene_types):
    color_map[gene_type] = colors[i % len(colors)]

# Color the cells in the 'Type' column
for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)
    gene_type = row[4]  # Assuming 'Type' is the 5th column
    if gene_type in color_map:
        cell = sheet.cell(row=sheet.max_row, column=5)  # Assuming 'Type' is the 5th column
        fill = PatternFill(start_color=color_map[gene_type], end_color=color_map[gene_type], fill_type="solid")
        cell.fill = fill

# Save the workbook as an XLSX file
xlsx_file = 'gene_data.xlsx'
workbook.save(xlsx_file)

print(f"Gene data exported to {xlsx_file}")
