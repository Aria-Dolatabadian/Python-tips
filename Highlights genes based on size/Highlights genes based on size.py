import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the CSV file
df = pd.read_csv('gene_data.csv')

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set up colors for highlighting
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

# Iterate over the rows and highlight based on the conditions
for row in dataframe_to_rows(df, index=False, header=True):
    if row[0] == 'Chromosome':  # Skip the header row
        sheet.append(row)
        continue
    sheet.append(row)
    if len(row) >= 4:
        try:
            start = int(row[2])  # Assuming 'Start' is the 3rd column
            end = int(row[3])  # Assuming 'End' is the 4th column
            diff = end - start
            if diff < 30000000:
                for cell in sheet.iter_cols(min_col=3, max_col=4, min_row=sheet.max_row, max_row=sheet.max_row):
                    for c in cell:
                        c.fill = red_fill
            elif diff > 50000000:
                for cell in sheet.iter_cols(min_col=3, max_col=4, min_row=sheet.max_row, max_row=sheet.max_row):
                    for c in cell:
                        c.fill = green_fill
            else:
                for cell in sheet.iter_cols(min_col=3, max_col=4, min_row=sheet.max_row, max_row=sheet.max_row):
                    for c in cell:
                        c.fill = blue_fill
        except ValueError:
            continue

# Save the workbook as an XLSX file
xlsx_file = 'highlighted_gene_data.xlsx'
workbook.save(xlsx_file)

print(f"Gene data with highlights exported to {xlsx_file}")
